# -*- coding: utf-8 -*-
"""生成公开数值分析表。运行: python docs/generate_balance_workbook.py"""
from __future__ import annotations

import math
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

OUT = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else (
    Path(__file__).resolve().parent / "数值分析_模块怪物波次.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 波 1–25 常规；波 26 在生成表时单独追加 Boss 行
WAVE_NORMAL = [4, 5, 6, 8, 8, 8, 11, 14, 18, 23, 26, 28, 30, 32, 36, 40, 44, 48, 52, 58, 62, 66, 72, 78, 88]
WAVE_SWARM = [0, 0, 0, 0, 4, 7, 11, 16, 23, 34, 40, 44, 48, 52, 60, 68, 76, 84, 92, 104, 112, 120, 130, 140, 160]
WAVE_TANK = [0, 0, 0, 0, 0, 0, 0, 1, 2, 4, 5, 6, 7, 8, 10, 11, 12, 14, 16, 18, 20, 22, 24, 26, 30]
WAVE_SAND = [0, 0, 0, 0, 0, 1, 2, 2, 2, 2, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 4, 4, 4, 4, 4]
WAVE_HP = [10, 10, 10, 10, 10, 25, 25, 25, 25, 25, 60, 60, 60, 60, 60, 120, 120, 120, 120, 120, 200, 200, 200, 200, 200]
WAVE_IV = [
    0.95, 0.88, 0.80, 0.72, 0.58, 0.72, 0.55, 0.42, 0.30, 0.18,
    0.38, 0.30, 0.24, 0.18, 0.12, 0.32, 0.26, 0.20, 0.15, 0.10,
    0.28, 0.22, 0.16, 0.12, 0.08,
]
BOSS_HP = 50_000
TOTAL_WAVES = 26

BASE_PRICE = {
    "激光": (10, "普通", 2.25, 1.0),
    "炸弹": (25, "普通", 2.25, 1.0),
    "雪花": (15, "普通", 2.25, 1.0),
    "火花": (15, "普通", 2.25, 1.0),
    "收束器": (15, "稀有", 2.35, 1.5),
    "矿机": (18, "稀有", 2.35, 1.5),
    "火焰增幅": (20, "稀有", 2.35, 1.5),
    "寒冰增幅": (20, "稀有", 2.35, 1.5),
    "传送门": (22, "稀有", 2.35, 1.5),
    "中续器": (22, "稀有", 2.35, 1.5),
    "加速器": (20, "稀有", 2.35, 1.5),
    "火附魔": (20, "稀有", 2.35, 1.5),
    "惊喜": (20, "稀有", 2.35, 1.5),
    "热浪": (25, "稀有", 2.35, 1.5),
    "冰霜冻结": (25, "稀有", 2.35, 1.5),
    "奥数飞弹": (22, "稀有", 2.35, 1.5),
    "激光炮": (30, "普通", 2.25, 1.0),
    "烈焰墙": (28, "稀有", 2.35, 1.5),
    "火焰祝福": (24, "稀有", 2.35, 1.5),
    "净化": (28, "稀有", 2.35, 1.5),
    "寒霜蘑菇": (28, "稀有", 2.35, 1.5),
    "黑洞": (45, "史诗", 2.60, 2.5),
    "聚变": (40, "史诗", 2.60, 2.5),
    "裂变": (40, "史诗", 2.60, 2.5),
    "分裂器": (55, "史诗", 2.60, 2.5),
}

ATTACK = {"激光", "炸弹", "雪花", "火花", "黑洞", "热浪", "冰霜冻结", "奥数飞弹", "激光炮", "烈焰墙"}
LEVEL_SCALE_FLAT = {"火焰增幅", "寒冰增幅", "火附魔", "惊喜"}

MODULE_ROWS: list = []


def add_mod(name, rarity, lv, dmg, cost, cap, rps, interval, special, dps, short, desc, note=""):
    """rps=发/秒（快模块）；interval=开火间隔秒（慢模块，与 UI「射速」同一概念）。"""
    MODULE_ROWS.append(
        [name, rarity, lv, dmg, cost, cap, rps, interval, special, dps, short, desc, note]
    )


def laser_dmg(lv: int) -> int:
    return round(5 * (1.8 ** (lv - 1)))


def bomb_dmg(lv: int) -> int:
    return laser_dmg(lv)


def spark_dmg(lv: int) -> int:
    return {1: 5, 2: 10, 3: 15, 4: 20, 5: 30}[lv]


def ice_dmg(lv: int) -> int:
    return 5 if lv == 1 else 10


def arcane_dmg(lv: int) -> int:
    return {1: 10, 2: 20, 3: 30, 4: 50, 5: 80}[lv]


def heatwave_dur(lv: int) -> float:
    return 2.0 if lv <= 3 else 3.0


def heatwave_cd(lv: int) -> float:
    return {1: 5.0, 2: 4.5, 3: 4.0, 4: 4.0, 5: 3.0}[lv]


def cannon_dmg(lv: int) -> int:
    return {1: 30, 2: 60, 3: 100, 4: 150, 5: 200}[lv]


def flame_amp(lv: int) -> int:
    return {1: 1, 2: 3, 3: 5, 4: 7, 5: 10}[lv]


def ice_amp_bonus(lv: int) -> int:
    return 5 * lv


def bh_radius(lv: int) -> float:
    return round(2.2 * (1 + 0.25 * (lv - 1)), 2)


def bh_duration(lv: int) -> float:
    return round(2.2 + 0.4 * (lv - 1), 1)


def bh_pull(lv: int) -> float:
    return round(3.5 + 0.8 * (lv - 1), 1)


def build_module_rows() -> None:
    MODULE_ROWS.clear()
    for lv in range(1, 6):
        dmg = laser_dmg(lv)
        add_mod(
            "激光", "普通", lv, dmg, 1, 5, 5, "", "",
            dmg * 5, "单体激光", "对最近敌人发射激光", "查理激光塔",
        )
    for lv in range(1, 6):
        dmg = bomb_dmg(lv)
        rps = 1.5
        radius = round(1.5 * (1 + 0.3 * (lv - 1)), 2)
        dps = round(dmg * rps, 1)  # 满能可持续：容5耗5→单发周期即间隔
        add_mod(
            "炸弹", "普通", lv, dmg, 5, 5, rps, "",
            f"AOE半径{radius}",
            dps, "最左AOE", "向最左敌人投掷炸弹爆炸", "",
        )
    for lv in range(1, 6):
        dmg = ice_dmg(lv)
        chill_s = 2 + (lv - 1)
        add_mod(
            "雪花", "普通", lv, dmg, 1, 5, 5, "",
            f"[寒冷]基础减速30% {chill_s}s", dmg * 5,
            "寒冷弹", "向最左敌人发射雪花飞弹，可以使敌人获得[寒冷]",
            "[寒冷]是状态；减速是其效果，可被寒冰增幅",
        )
    for lv in range(1, 6):
        dmg = spark_dmg(lv)
        burn = round(2 + 0.5 * (lv - 1), 1)
        add_mod(
            "火花", "普通", lv, dmg, 1, 5, 5, "",
            f"灼烧{burn}s", dmg * 5,
            "灼烧弹", "向最左敌人发射火花飞弹，可以使敌人获得[灼烧]", "",
        )
    for lv in range(1, 6):
        dmg = cannon_dmg(lv)
        rps = 1
        add_mod(
            "激光炮", "普通", lv, dmg, 5, 5, rps, "", "厚激光",
            dmg * rps, "高伤激光", "缓慢对最近敌人发射更大的激光", "",
        )
    for lv in range(1, 6):
        dmg = arcane_dmg(lv)
        add_mod(
            "奥数飞弹", "稀有", lv, dmg, 1, 5, 5, "", "索敌最右",
            dmg * 5, "紫色飞弹", "紫色索敌飞弹，锁定最右敌人", "",
        )
    for lv in range(1, 6):
        dur = heatwave_dur(lv)
        interval = heatwave_cd(lv)
        add_mod(
            "热浪", "稀有", lv, "", 20, 20, "", interval, f"全屏灼烧{dur}s",
            "", "全屏灼烧", "全屏敌人获得[灼烧]", "射速以间隔秒显示",
        )
    for lv in range(1, 6):
        dur = heatwave_dur(lv)
        interval = heatwave_cd(lv)
        add_mod(
            "冰霜冻结", "稀有", lv, "", 20, 20, "", interval,
            f"全屏[寒冷] {dur}s（基础减速30%）",
            "", "全屏寒冷", "全屏敌人获得[寒冷]",
            "[寒冷]≠减速同义词；减速为基础效果，可被寒冰增幅",
        )
    for lv in range(1, 6):
        dmg = {1: 5, 2: 10, 3: 15, 4: 29, 5: 50}[lv]
        burn = {1: 2, 2: 2, 3: 3, 4: 4, 5: 5}[lv]
        add_mod(
            "烈焰墙", "稀有", lv, dmg, "", 30, "", "",
            f"穿墙伤+灼烧{burn}s", "", "路径火墙", "能量球穿过时造成伤害并挂烧", "",
        )
    for lv in range(1, 6):
        add_mod(
            "火焰增幅", "稀有", lv, "", "", "", "", "",
            f"灼烧+{flame_amp(lv)}/0.5s", "", "灼烧增幅", "场上被动提高灼烧跳动伤害", "",
        )
    for lv in range(1, 6):
        bonus = ice_amp_bonus(lv)
        add_mod(
            "寒冰增幅", "稀有", lv, "", "", "", "", "",
            f"寒冷减速+{bonus}%（总上限70%）", "", "寒冷增幅",
            "场上被动：寒冷造成的减速每级+5%，总减速上限70%",
            "[寒冷]状态的减速效果增幅",
        )
    for lv in range(1, 4):
        gold = {1: 5, 2: 10, 3: 20}[lv]
        add_mod(
            "矿机", "稀有", lv, "", 10, 10, "", 3, f"{gold}金/次",
            "", "产金", "耗能产金", "最高Lv3",
        )
    for name, short, desc in [
        ("收束器", "直角改向", "将光球沿直角改向"),
        ("传送门", "成对传送", "成对传送并保持方向，最多2座"),
        ("中续器", "汲能续寿", "穿过汲能，满后刷新球寿命"),
        ("加速器", "球速×1.5", "未加速球速度×1.5（每球一次）"),
        ("火附魔", "灼烧附魔格", "种子格写入灼烧附魔"),
        ("惊喜", "随机附魔格", "种子格写入随机附魔"),
        ("火焰祝福", "道具", "手牌火焰祝福道具"),
        ("净化", "道具", "净化诅咒/锁定"),
        ("寒霜蘑菇", "道具", "紧急寒冷道具"),
    ]:
        rarity = BASE_PRICE[name][1]
        add_mod(name, rarity, 1, "", "", "", "", "", "", "", short, desc, "")
    for lv in range(1, 6):
        add_mod(
            "黑洞", "史诗", lv, "", 5, 5, "", 6,
            f"半径{bh_radius(lv)} 持续{bh_duration(lv)}s 吸力{bh_pull(lv)}",
            "", "吸引聚怪",
            "每6秒耗5能投掷黑洞吸引敌人（半径/时长/吸力随等级提升）", "",
        )
    for name, short, desc in [
        ("聚变", "五合一", "5球合成1球"),
        ("裂变", "一变五", "≥5能射出5颗默认球"),
        ("分裂器", "一分二", "T形分裂，寿命减半"),
    ]:
        rarity = BASE_PRICE[name][1]
        add_mod(name, rarity, 1, "", "", "", "", "", "", "", short, desc, "")


def stage(w: int) -> int:
    return (max(1, w) - 1) // 5


def round5(v: float) -> int:
    if v <= 0:
        return 0
    return max(5, int(round(v / 5.0) * 5))


def gold_budget(w: int) -> int:
    w = max(1, min(TOTAL_WAVES, w))
    return max(20, round5(round(18 * (1.205 ** (w - 1)))))


def draft_tags(w: int) -> str:
    """与 ModuleUnlockDirector / EmitterUpgradeDirector / BlessingCurse 节奏对齐。"""
    unlock = []
    if w < TOTAL_WAVES:
        if (w < 15 and w % 3 == 0) or (w >= 15 and w % 2 == 0):
            unlock.append("模块")
        if w % 5 == 0:
            unlock.append("熔炉")
            unlock.append("祝福束缚")
    return "+".join(unlock)


def shop_price(name: str, lv: int, wave: int = 1) -> int:
    base, _, lv_exp, rarity = BASE_PRICE[name]
    if name in ATTACK:
        price = base * (lv_exp ** (lv - 1)) * rarity
    elif name == "矿机":
        price = base * (1 + 0.35 * (lv - 1)) * rarity
    elif name in LEVEL_SCALE_FLAT:
        price = base * (1 + 0.4 * (lv - 1)) * rarity
    else:
        price = base * rarity
    return round5(round(price))


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, min_w=8, max_w=48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = min_w
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = length


def write_table(ws, start_row: int, headers: list, rows: list) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(start_row, i, h)
    style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for i, v in enumerate(row, 1):
            cell = ws.cell(r, i, v)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
    return r


def main() -> None:
    build_module_rows()
    wb = Workbook()

    ws = wb.active
    ws.title = "00_说明"
    notes = [
        ["GMTK2026 塔防 — 数值分析表（由 docs/generate_balance_workbook.py 生成）"],
        ["配套说明", "docs/数值平衡表.md / docs/游戏策划案.md"],
        ["真相优先级", "Assets/Scripts 代码 > 本表 > 旧文档"],
        [""],
        ["工作表"],
        ["01_全局", "开局与常量"],
        ["02_模块总表", "★各模块每级数值/小描述/描述/满能DPS（方便手改）"],
        ["03_熔炉", "四维档位"],
        ["04_波次", "26 波配额/HP/间隔/金币预算（波26=Boss）"],
        ["05_商店价", "各模块等级标价"],
        [""],
        ["近期对齐"],
        ["总波数", "26；波26仅 Boss HP=50000；移速0.375；Boss漏家即败"],
        ["补沙", "仅沙buff击杀；清波不补沙"],
        ["射速", "统一开火间隔概念；无独立冷却附魔"],
        ["窗口化", "1440×900；全屏=桌面分辨率"],
        ["奥数飞弹", "稀有；能耗1/容5/射速5；伤10/20/30/50/80；最右索敌"],
    ]
    for r, row in enumerate(notes, 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            if r == 1 or (c == 1 and v):
                cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 70

    ws = wb.create_sheet("01_全局")
    write_table(
        ws, 1, ["项", "值", "备注"],
        [
            ["起始金币", 50, ""],
            ["总波数", TOTAL_WAVES, "1–25常规 + 26 Boss"],
            ["Boss HP", BOSS_HP, "波26；漏家立刻败北"],
            ["Boss 移速", 0.375, "坦克0.75的一半；单位格/秒"],
            ["红/黄/蓝移速", "1.5 / 3.0 / 0.75", "紫拆·金盾同蓝"],
            ["开局沙 ms", 100_000, "沙尽即败"],
            ["抽沙", "1000 ms/秒", "仅战斗"],
            ["补沙", "仅沙buff击杀", "清波不补沙"],
            ["手牌/商店", "8 / 6", ""],
            ["商店池上限", 12, "开局：激光+收束+雪花+火花"],
            ["球上限", 40, ""],
            ["球寿命档", "12/20/32/50 s", "熔炉寿命维"],
            ["窗口化", "1440×900", "全屏=桌面分辨率"],
            ["分解返还", 0.30, "invested×30%"],
            ["扩展费用", "100 / 300", "3→5 / 5→7"],
            ["刷新费", "5×((shopLv+1)/2)", "商店 Lv1–2→5，Lv3–4→10…"],
            ["商店等级", "(wave-1)//5+1", "每5波升1级；奇级单档/偶级双档"],
            ["模块解锁", "前15每3波，后每2波", "3/6/9/12/16/18/20/22/24"],
            ["熔炉/祝福", "每5波至25", "5/10/15/20/25"],
            ["胜负", "返回主菜单", "ResultOverlay"],
        ],
    )
    autosize(ws)

    ws = wb.create_sheet("02_模块总表")
    write_table(
        ws, 1,
        ["模块", "稀有度", "Lv", "伤害", "能耗", "容量", "射速发/秒", "间隔秒",
         "特殊效果", "满能DPS", "小描述", "描述", "备注"],
        MODULE_ROWS,
    )
    note = ws.cell(
        2 + len(MODULE_ROWS), 1,
        "射速与间隔为同一概念（UI统称射速）。满能DPS≈伤×射速（可持续）。热浪/冰霜/矿机等控场经济类为空，请手改后同步 ModuleCatalog。",
    )
    note.fill = NOTE_FILL
    autosize(ws)

    ws = wb.create_sheet("03_熔炉")
    caps = [2000, 1400, 1050, 800]
    speeds = [4.0, 5.5, 7.0, 8.5]
    mass = [1, 1.5, 2, 2.5]
    life = [12, 20, 32, 50]
    rows = []
    for i in range(4):
        rate = 1000 / caps[i]
        rows.append([i, caps[i], round(rate, 2), speeds[i], mass[i], life[i], round(rate * mass[i], 2)])
    write_table(ws, 1, ["档", "容量ms", "出球/秒", "球速", "质量", "寿命s", "能量/秒"], rows)
    autosize(ws)

    ws = wb.create_sheet("04_波次")
    rows = []
    prev_hp = None
    for w in range(1, 26):
        i = w - 1
        n, s, t = WAVE_NORMAL[i], WAVE_SWARM[i], WAVE_TANK[i]
        hp = WAVE_HP[i]
        total = n * hp + s * math.ceil(hp * 0.5) + t * hp * 4
        growth = "" if prev_hp is None else round(total / prev_hp - 1, 3)
        prev_hp = total
        rows.append([
            w, stage(w), n, s, t, WAVE_SAND[i], WAVE_IV[i], hp, total, growth,
            gold_budget(w), draft_tags(w),
        ])
    rows.append([
        26, stage(26), 0, 0, 0, 0, "", f"Boss {BOSS_HP}", BOSS_HP,
        "" if prev_hp is None else round(BOSS_HP / prev_hp - 1, 3),
        gold_budget(26), draft_tags(26),
    ])
    write_table(
        ws, 1,
        ["波", "stage", "红", "黄", "蓝", "沙buff基", "间隔s", "红HP", "总HP", "总HP环比", "波金币", "草稿"],
        rows,
    )
    autosize(ws)

    ws = wb.create_sheet("05_商店价")
    headers = ["模块", "稀有度", "基础价", "Lv1", "Lv2", "Lv3", "Lv4", "Lv5"]
    rows = []
    for name, (base, rarity, _, _) in BASE_PRICE.items():
        if name == "矿机":
            max_lv = 3
        elif name in ("火附魔", "惊喜"):
            max_lv = 4
        elif name in ("火焰祝福", "净化", "寒霜蘑菇"):
            max_lv = 1
        elif name in ATTACK or name in LEVEL_SCALE_FLAT or name == "烈焰墙":
            max_lv = 5
        else:
            max_lv = 1
        cells = [name, rarity, base]
        for lv in range(1, 6):
            cells.append(shop_price(name, lv) if lv <= max_lv else "")
        rows.append(cells)
    write_table(ws, 1, headers, rows)
    autosize(ws)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
